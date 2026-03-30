import glob
import json
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Serialization helpers ──────────────────────────────────────────────────────

def _val(v):
    """Return a clean string; treat None and whitespace-only as empty."""
    return str(v).strip() if v is not None else ""


def _serialize_ecg(rows):
    """'TIME TYPE RHYTHM NOTES' per row; rows separated by a single space."""
    parts = []
    for row in rows:
        t     = _val(row.get("Time"))
        typ   = _val(row.get("Type"))   or "NULL"
        rhy   = _val(row.get("Rhythm")) or "NULL"
        notes = _val(row.get("Notes"))  or "NULL"
        parts.append(" ".join(x for x in [t, typ, rhy, notes] if x))
    return " ".join(parts)


def _serialize_flow_chart(rows):
    """'TIME TREATMENT DESCRIPTION' per row; rows separated by ' | '."""
    parts = []
    for row in rows:
        t    = _val(row.get("Time"))
        trt  = _val(row.get("Treatment"))
        desc = _val(row.get("Description"))
        entry = " ".join(x for x in [t, trt, desc] if x)
        if entry:
            parts.append(entry)
    return " | ".join(parts)


def _serialize_assessments(section):
    """Multi-line 'Assessment time: KEY\\nFIELD: VALUE\\n...' per time block."""
    if not section or not isinstance(section, dict):
        return ""
    lines = []
    for time_key, fields in section.items():
        lines.append(f"Assessment time: {time_key}")
        if isinstance(fields, dict):
            for field, value in fields.items():
                if field == "Assessment Time":
                    continue
                v = _val(value)
                if v:
                    lines.append(f"{field}: {v}")
    return "\n".join(lines)


def _all_timed(rows, field):
    """All rows serialized as 'TIME val; TIME val; ...'"""
    parts = []
    for row in rows:
        t = _val(row.get("Time"))
        v = _val(row.get(field)) or "NULL"
        parts.append(f"{t} {v}".strip())
    return "; ".join(parts)


def _concat_newline(rows, field):
    """Join the field value from every row with '\\n'."""
    values = [_val(row.get(field)) for row in rows if _val(row.get(field))]
    return "\n".join(values)


# ── Derived Flow Chart values ──────────────────────────────────────────────────

_ACCESS_RE = re.compile(
    r'\b(IV|IO|Intravenous|Intraosseous|Vascular Access|PIV|EZ-IO|Central Line)\b',
    re.IGNORECASE,
)
_ROUTE_RE  = re.compile(r'Route:\s*([^;(]+)', re.IGNORECASE)
_AIRWAY_RE = re.compile(
    r'\b(Intubat|Supraglottic|LMA|King|iGel|i-Gel|Cricotr|RSI|ETT|Trach)\b',
    re.IGNORECASE,
)
_EPI_RE    = re.compile(r'epinephrine', re.IGNORECASE)


def _derive_flow_chart_values(rows):
    """
    Scan Flow Chart rows and return a dict with three derived fields:
        first_access  – route/type of the first IV/IO access entry
        epi_count     – number of Epinephrine administration entries
        final_airway  – device name of the last airway-management entry
    All values are strings.
    """
    first_access = ""
    last_airway  = ""
    epi_count    = 0

    for row in rows:
        treatment   = _val(row.get("Treatment"))
        description = _val(row.get("Description"))

        if not first_access and _ACCESS_RE.search(treatment):
            route_match = _ROUTE_RE.search(description)
            if route_match:
                first_access = route_match.group(1).strip()
            elif " - " in treatment:
                first_access = treatment.split(" - ")[-1].strip()
            else:
                first_access = treatment

        if _EPI_RE.search(treatment):
            epi_count += 1

        if _AIRWAY_RE.search(treatment):
            last_airway = treatment

    if " - " in last_airway:
        last_airway = last_airway.split(" - ")[-1].strip()

    return {
        "first_access": first_access,
        "epi_count":    str(epi_count) if epi_count else "",
        "final_airway": last_airway,
    }


# ── Single-value resolver ──────────────────────────────────────────────────────

def get_value(tables, mapping, derived_cache):
    """
    Resolve one WORD_TO_DATA_MAP entry to a plain string.

    Parameters
    ----------
    tables        : dict – the 'tables' dict from a parsed JSON file
    mapping       : tuple | None – a WORD_TO_DATA_MAP value
    derived_cache : dict – pre-computed derived values for the current incident
    """
    if mapping is None:
        return ""

    if mapping[0] == "__derived__":
        return derived_cache.get(mapping[1], "")

    section  = mapping[0]
    field    = mapping[1]
    modifier = mapping[2] if len(mapping) > 2 else None

    section_data = tables.get(section)
    if section_data is None:
        return ""

    # Entire-section serializers (field is None)
    if field is None:
        if section == "ECG":
            return _serialize_ecg(section_data) if isinstance(section_data, list) else ""
        if section == "Flow Chart":
            return _serialize_flow_chart(section_data) if isinstance(section_data, list) else ""
        if section == "Assessments":
            return _serialize_assessments(section_data)
        return ""

    # Split fields: "val1 unit1 - val2 unit2" → take before or after " - "
    if modifier in ("before_sep", "after_sep") and isinstance(section_data, dict):
        raw = _val(section_data.get(field, ""))
        parts = raw.split(" - ", 1)
        if modifier == "before_sep":
            return parts[0].strip()
        return parts[1].strip() if len(parts) > 1 else ""

    # List-based sections
    if isinstance(section_data, list):
        if modifier == "all_timed":
            return _all_timed(section_data, field)
        if modifier == "concat_newline":
            return _concat_newline(section_data, field)
        if isinstance(modifier, int):
            return _val(section_data[modifier].get(field)) if modifier < len(section_data) else ""
        return _val(section_data[0].get(field)) if section_data else ""

    # Standard key-value dict sections
    if isinstance(section_data, dict):
        return _val(section_data.get(field, ""))

    return ""


# ── Row extraction ─────────────────────────────────────────────────────────────

def extract_row(data):
    """Build a flat ordered dict for one incident, ready to become an Excel row."""
    tables          = data.get("tables", {})
    flow_chart_rows = tables.get("Flow Chart", [])
    derived         = _derive_flow_chart_values(
        flow_chart_rows if isinstance(flow_chart_rows, list) else []
    )

    row = {
        "Incident Number": _val(data.get("incident_number")),
        "Incident Date":   _val(data.get("incident_date")),
    }
    for col_name, mapping in WORD_TO_DATA_MAP.items():
        row[col_name] = get_value(tables, mapping, derived)

    return row


# ── Excel export ───────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def export_to_excel(rows, output_path):
    """Write a list of row-dicts (from extract_row) to an .xlsx file."""
    if not rows:
        print("No rows to export.")
        return

    headers = list(rows[0].keys())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incidents"

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Data rows
    for row_idx, row_dict in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get(header, ""))

    # Column widths: fit to content, capped at 60 characters
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, len(rows) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value or ""
            first_line = str(cell_val).split("\n")[0]
            max_len = max(max_len, min(len(first_line), 60))
        ws.column_dimensions[col_letter].width = max_len + 2

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"Saved {len(rows)} rows -> {output_path}")


# Mapping from Word document column headers to their location in the parsed JSON data.
#
# Tuple formats:
#   (section, field)
#       Regular key-value section lookup: data["tables"][section][field]
#       field=None → serialize the entire section as a string (ECG, Flow Chart, Assessments)
#
#   (section, field, "all_timed")
#       List-based section where ALL rows are needed.
#       Serialized as "TIME val; TIME val; …" using the "Time" key of each row.
#       Used for Vital Signs and Vitals Calculations.
#
#   (section, field, "concat_newline")
#       List-based section where ALL rows are needed.
#       Values for `field` across every row are joined with "\n".
#       Used for Specialty Patient - Advanced Airway fields.
#
#   (section, field, int)
#       List-based section: pick the row at the given integer index.
#       Used for Consumables (index 0, 1, 2).
#
#   ("__derived__", key)
#       Value is computed from Flow Chart entries at export time.
#       Supported keys:
#           "first_access"  – treatment type of the first IV/IO access entry
#           "epi_count"     – number of Epinephrine administrations
#           "final_airway"  – device name of the last airway-management entry
#
#   None
#       No equivalent field exists in the extracted data (cell will be empty).

WORD_TO_DATA_MAP = {
    # ── Patient Information ──────────────────────────────────────────────────
    "Patient Information - Last":                              ("Patient Information", "Last"),
    "Patient Information - First":                             ("Patient Information", "First"),
    "Patient Information - Middle":                            ("Patient Information", "Middle"),
    "Patient Information - Name Suffix":                       ("Patient Information", "Name Suffix"),
    "Patient Information - Sex":                               ("Patient Information", "Sex"),
    "Patient Information - Gender":                            ("Patient Information", "Gender"),
    "Patient Information - DOB":                               ("Patient Information", "DOB"),
    "Patient Information - Age":                               ("Patient Information", "Age"),
    "Patient Information - Weight-lbs":                        ("Patient Information", "Weight", "before_sep"),
    "Patient Information - Weight-kg":                         ("Patient Information", "Weight", "after_sep"),
    "Patient Information - Height-ft":                         ("Patient Information", "Height", "before_sep"),
    "Patient Information - Height- cm":                        ("Patient Information", "Height", "after_sep"),
    "Patient Information - Pedi Color":                        ("Patient Information", "Pedi Color"),
    "Patient Information - SSN":                               ("Patient Information", "SSN"),
    "Patient Information - Advance Directives":                ("Patient Information", "Advance Directives"),
    "Patient Information - Resident Status":                   ("Patient Information", "Resident Status"),
    "Patient Information - Patient Resides in Service Area":   ("Patient Information", "Patient Resides in Service Area"),
    "Patient Information - Temporary Residence Type":          ("Patient Information", "Temporary Residence Type"),
    "Patient Information - Address":                           ("Patient Information", "Address"),
    "Patient Information - Address 2":                         ("Patient Information", "Address 2"),
    "Patient Information - City":                              ("Patient Information", "City"),
    "Patient Information - State":                             ("Patient Information", "State"),
    "Patient Information - Zip":                               ("Patient Information", "Zip"),
    "Patient Information - Country":                           ("Patient Information", "Country"),
    "Patient Information - Tel":                               ("Patient Information", "Tel"),
    "Patient Information - Physician":                         ("Patient Information", "Physician"),
    "Patient Information - Phys. Tel":                         ("Patient Information", "Phys. Tel"),
    "Patient Information - Ethnicity":                         ("Patient Information", "Ethnicity"),
    "Patient Information - Race":                              ("Patient Information", "Race"),

    # ── Clinical Impression ──────────────────────────────────────────────────
    "Clinical Impression - Primary Impression":                ("Clinical Impression", "Primary Impression"),
    "Clinical Impression - Secondary Impression":              ("Clinical Impression", "Secondary Impression"),
    "Clinical Impression - Protocols Used":                    ("Clinical Impression", "Protocols Used"),
    "Clinical Impression - Local Protocol Provided Care Level":("Clinical Impression", "Local Protocol Provided Care Level"),
    "Clinical Impression - Anatomic Position":                 ("Clinical Impression", "Anatomic Position"),
    "Clinical Impression - Onset Time":                        ("Clinical Impression", "Onset Time"),
    "Clinical Impression - Last Known Well":                   ("Clinical Impression", "Last Known Well"),
    "Clinical Impression - Chief Complaint":                   ("Clinical Impression", "Chief Complaint"),
    "Clinical Impression - Duration (Chief Complaint)":        ("Clinical Impression", "Duration"),
    "Clinical Impression - Units (Chief Complaint Duration)":  ("Clinical Impression", "Duration Units"),
    "Clinical Impression - Secondary Complaint":               ("Clinical Impression", "Secondary Complaint"),
    "Clinical Impression - Duration (Secondary Complaint)":    ("Clinical Impression", "Secondary Duration"),
    "Clinical Impression - Units (Secondary Complaint Duration)": ("Clinical Impression", "Secondary Duration Units"),
    "Clinical Impression - Patient's Level of Distress":       ("Clinical Impression", "Patient Level of Distress"),
    "Clinical Impression - Signs & Symptoms":                  ("Clinical Impression", "Signs & Symptoms"),
    "Clinical Impression - Injury":                            ("Clinical Impression", "Injury"),
    "Clinical Impression - Additional Injury":                 ("Clinical Impression", "Additional Injury"),
    "Clinical Impression - Mechanism of Injury":               ("Clinical Impression", "Mechanism of Injury"),
    "Clinical Impression - Medical/Trauma":                    ("Clinical Impression", "Medical/Trauma"),
    "Clinical Impression - Barriers of Care":                  ("Clinical Impression", "Barriers of Care"),
    "Clinical Impression - Alcohol/Drugs":                     ("Clinical Impression", "Alcohol/Drugs"),
    "Clinical Impression - Pregnancy":                         ("Clinical Impression", "Pregnancy"),
    "Clinical Impression - Initial Patient Acuity":            ("Clinical Impression", "Initial Patient Acuity"),
    "Clinical Impression - Final Patient Acuity":              ("Clinical Impression", "Final Patient Acuity"),
    "Clinical Impression - Patient Activity":                  ("Clinical Impression", "Patient Activity"),

    # ── Medications / Allergies / History / Immunizations ───────────────────
    "Medications":  ("Medications/Allergies/History/Immunizations", "Medications"),
    "Allergies":    ("Medications/Allergies/History/Immunizations", "Allergies"),
    "History":      ("Medications/Allergies/History/Immunizations", "History"),
    # Word uses "Immunization" (no trailing s)
    "Immunization": ("Medications/Allergies/History/Immunizations", "Immunizations"),

    # ── Vital Signs (all rows, prefixed with their timestamp) ────────────────
    "Vital Signs - AVPU":  ("Vital Signs", "AVPU",  "all_timed"),
    "Vital Signs - Side":  ("Vital Signs", "Side",  "all_timed"),
    "Vital Signs - POS":   ("Vital Signs", "POS",   "all_timed"),
    "Vital Signs - BP":    ("Vital Signs", "BP",    "all_timed"),
    "Vital Signs - Pulse": ("Vital Signs", "Pulse", "all_timed"),
    "Vital Signs - RR":    ("Vital Signs", "RR",    "all_timed"),
    "Vital Signs - SPO2":  ("Vital Signs", "SPO2",  "all_timed"),
    "Vital Signs - ETCO2": ("Vital Signs", "ETCO2", "all_timed"),
    "Vital Signs - CO":    ("Vital Signs", "CO",    "all_timed"),
    "Vital Signs - BG":    ("Vital Signs", "BG",    "all_timed"),
    "Vital Signs - Temp":  ("Vital Signs", "Temp",  "all_timed"),
    "Vital Signs - Pain":  ("Vital Signs", "Pain",  "all_timed"),

    # ── Vitals Calculations (all rows, prefixed with their timestamp) ─────────
    "Vitals Calculations - GCS(E+V+M)/Qualifiers": ("Vitals Calculations", "GCS(E+V+M)/Qualifiers", "all_timed"),
    "Vitals Calculations - RASS":                  ("Vitals Calculations", "RASS",        "all_timed"),
    "Vitals Calculations - BARS":                  ("Vitals Calculations", "BARS",        "all_timed"),
    "Vitals Calculations - RTS":                   ("Vitals Calculations", "RTS",         "all_timed"),
    "Vitals Calculations - PTS":                   ("Vitals Calculations", "PTS",         "all_timed"),
    "Vitals Calculations - MAP":                   ("Vitals Calculations", "MAP",         "all_timed"),
    "Vitals Calculations - Shock Index":            ("Vitals Calculations", "Shock Index", "all_timed"),

    # ── ECG & Flow Chart (entire list serialized into one cell) ──────────────
    "ECG":        ("ECG",        None),
    "Flow Chart":  ("Flow Chart", None),

    # ── Derived from Flow Chart entries ──────────────────────────────────────
    "First access":                          ("__derived__", "first_access"),
    "Number of epinephrine administrations": ("__derived__", "epi_count"),
    "Final airway device":                   ("__derived__", "final_airway"),

    # ── Assessments (entire dict serialized into one cell) ───────────────────
    "Assessments": ("Assessments", None),

    # ── Narrative ────────────────────────────────────────────────────────────
    "Narrative": ("Narrative", "Narrative"),

    # ── Specialty Patient - Advanced Airway (all rows joined with newline) ───
    "Airway":                    ("Specialty Patient - Advanced Airway", "Airway",                    "concat_newline"),
    "Indications":               ("Specialty Patient - Advanced Airway", "Indications",               "concat_newline"),
    "Monitoring Devices":        ("Specialty Patient - Advanced Airway", "Monitoring Devices",        "concat_newline"),
    "Rescue Devices":            ("Specialty Patient - Advanced Airway", "Rescue Devices",            "concat_newline"),
    "Reasons Failed Intubation": ("Specialty Patient - Advanced Airway", "Reasons Failed Intubation", "concat_newline"),

    # ── CPR ──────────────────────────────────────────────────────────────────
    "CPR - Cardiac Arrest":              ("Specialty Patient - CPR", "Cardiac Arrest"),
    "CPR - Cardiac Arrest Etiology":     ("Specialty Patient - CPR", "Cardiac Arrest Etiology"),
    "CPR - Estimated Time of Arrest":    ("Specialty Patient - CPR", "Estimated Time of Arrest"),
    "CPR - Est Time Collapse to 911":    ("Specialty Patient - CPR", "Est Time Collapse to 911"),
    "CPR - Est Time Collapse to CPR":    ("Specialty Patient - CPR", "Est Time Collapse to CPR"),
    "CPR - Arrest Witnessed By":         ("Specialty Patient - CPR", "Arrest Witnessed By"),
    "CPR - CPR Initiated By":            ("Specialty Patient - CPR", "CPR Initiated By"),
    "CPR - Time 1st CPR":                ("Specialty Patient - CPR", "Time 1st CPR"),
    "CPR - CPR Feedback":                ("Specialty Patient - CPR", "CPR Feedback"),
    "CPR - ITD Used":                    ("Specialty Patient - CPR", "ITD Used"),
    "CPR - Applied AED":                 ("Specialty Patient - CPR", "Applied AED"),
    "CPR - Applied By":                  ("Specialty Patient - CPR", "Applied By"),
    "CPR - Defibrillated":               ("Specialty Patient - CPR", "Defibrillated"),
    "CPR - CPR Type":                    ("Specialty Patient - CPR", "CPR Type"),
    "CPR - Prearrival CPR Instructions": ("Specialty Patient - CPR", "Prearrival CPR Instructions"),
    "CPR - First Defibrillated By":      ("Specialty Patient - CPR", "First Defibrillated By"),
    "CPR - Time of First Defib":         ("Specialty Patient - CPR", "Time of First Defib"),
    "CPR - Initial ECG Rhythm":          ("Specialty Patient - CPR", "Initial ECG Rhythm"),
    "CPR - Rhythm at Destination":       ("Specialty Patient - CPR", "Rhythm at Destination"),
    "CPR - Hypothermia":                 ("Specialty Patient - CPR", "Hypothermia"),
    "CPR - End of Event":                ("Specialty Patient - CPR", "End of Event"),
    "CPR - ROSC":                        ("Specialty Patient - CPR", "ROSC"),
    "CPR - ROSC Time":                   ("Specialty Patient - CPR", "ROSC Time"),
    "CPR - ROSC Occurred":               ("Specialty Patient - CPR", "ROSC Occurred"),
    "CPR - Resuscitation Discontinued":  ("Specialty Patient - CPR", "Resuscitation Discontinued"),
    "CPR - Discontinued Reason":         ("Specialty Patient - CPR", "Discontinued Reason"),
    "CPR - Resuscitation":               ("Specialty Patient - CPR", "Resuscitation"),
    # "In Field Pronouncement" fields live inside the CPR section in the data
    "In Field Pronouncement - Expired":   ("Specialty Patient - CPR", "Expired"),
    "In Field Pronouncement - Time":      ("Specialty Patient - CPR", "Time"),
    "In Field Pronouncement - Date":      ("Specialty Patient - CPR", "Date"),
    "In Field Pronouncement - Physician": ("Specialty Patient - CPR", "Physician"),

    # ── Motor Vehicle Collision ───────────────────────────────────────────────
    "Motor Vehicle Collision - Patient Injured":        ("Specialty Patient - Motor Vehicle Collision", "Patient Injured"),
    "Motor Vehicle Collision - Vehicle Type":           ("Specialty Patient - Motor Vehicle Collision", "Vehicle Type"),
    "Motor Vehicle Collision - Position In Vehicle":    ("Specialty Patient - Motor Vehicle Collision", "Position In Vehicle"),
    "Motor Vehicle Collision - Seat Row":               ("Specialty Patient - Motor Vehicle Collision", "Seat Row"),
    "Motor Vehicle Collision - Number Of Vehicles":     ("Specialty Patient - Motor Vehicle Collision", "Number Of Vehicles"),
    "Motor Vehicle Collision - Weather":                ("Specialty Patient - Motor Vehicle Collision", "Weather"),
    "Motor Vehicle Collision - Extrication Required":   ("Specialty Patient - Motor Vehicle Collision", "Extrication Required"),
    "Motor Vehicle Collision - Estimated Speed":        ("Specialty Patient - Motor Vehicle Collision", "Estimated Speed"),
    "Motor Vehicle Collision - Exterior Damage":        ("Specialty Patient - Motor Vehicle Collision", "Exterior Damage"),
    "Motor Vehicle Collision - Law Enforcement Case #": ("Specialty Patient - Motor Vehicle Collision", "Law Enforcement Case #"),
    "Motor Vehicle Collision - Collision Indicators":   ("Specialty Patient - Motor Vehicle Collision", "Collision Indicators"),
    "Motor Vehicle Collision - Damage Location":        ("Specialty Patient - Motor Vehicle Collision", "Damage Location"),
    "Motor Vehicle Collision - Airbag Deployment":      ("Specialty Patient - Motor Vehicle Collision", "Airbag Deployment"),
    "Motor Vehicle Collision - Safety Devices":         ("Specialty Patient - Motor Vehicle Collision", "Safety Devices"),
    "Motor Vehicle Collision - Extrication Comments":   ("Specialty Patient - Motor Vehicle Collision", "Extrication Comments"),
    "Motor Vehicle Collision - Extrication Time":       ("Specialty Patient - Motor Vehicle Collision", "Extrication Time"),

    # ── Trauma Criteria ───────────────────────────────────────────────────────
    "Trauma Criteria - Anatomic":             ("Specialty Patient - Trauma Criteria", "Anatomic"),
    "Trauma Criteria - Physiologic":          ("Specialty Patient - Trauma Criteria", "Physiologic"),
    "Trauma Criteria - Mechanical":           ("Specialty Patient - Trauma Criteria", "Mechanical"),
    "Trauma Criteria - Other Conditions":     ("Specialty Patient - Trauma Criteria", "Other Conditions"),
    "Trauma Criteria - Trauma Activation":    ("Specialty Patient - Trauma Criteria", "Trauma Activation"),
    "Trauma Criteria - Time":                 ("Specialty Patient - Trauma Criteria", "Time"),
    "Trauma Criteria - Date":                 ("Specialty Patient - Trauma Criteria", "Date"),
    "Trauma Criteria - Trauma level":         ("Specialty Patient - Trauma Criteria", "Trauma level"),
    "Trauma Criteria - Reason Not Activated": ("Specialty Patient - Trauma Criteria", "Reason Not Activated"),

    # ── CDC 2011 Trauma Criteria ──────────────────────────────────────────────
    "CDC 2011 Trauma Criteria - Vital Signs":            ("Specialty Patient - CDC 2011 Trauma Criteria", "Vital Signs"),
    "CDC 2011 Trauma Criteria - Anatomy of Injury":      ("Specialty Patient - CDC 2011 Trauma Criteria", "Anatomy of Injury"),
    "CDC 2011 Trauma Criteria - Mechanism of Injury":    ("Specialty Patient - CDC 2011 Trauma Criteria", "Mechanism of Injury"),
    "CDC 2011 Trauma Criteria - Special Considerations": ("Specialty Patient - CDC 2011 Trauma Criteria", "Special Considerations"),
    "CDC 2011 Trauma Criteria - Trauma Activation":      ("Specialty Patient - CDC 2011 Trauma Criteria", "Trauma Activation"),
    "CDC 2011 Trauma Criteria - Time":                   ("Specialty Patient - CDC 2011 Trauma Criteria", "Time"),
    "CDC 2011 Trauma Criteria - Date":                   ("Specialty Patient - CDC 2011 Trauma Criteria", "Date"),
    "CDC 2011 Trauma Criteria - level":                  ("Specialty Patient - CDC 2011 Trauma Criteria", "level"),
    "CDC 2011 Trauma Criteria - Reason Not Activated":   ("Specialty Patient - CDC 2011 Trauma Criteria", "Reason Not Activated"),

    # ── Spinal Immobilization (all rows joined with newline) ─────────────────
    "Spinal Immobilization - Immobilization Recommended?":          ("Specialty Patient - Spinal Immobilization", "Immobilization Recommended?",          "concat_newline"),
    "Spinal Immobilization - Altered Mental Status":                ("Specialty Patient - Spinal Immobilization", "Altered Mental Status",                "concat_newline"),
    "Spinal Immobilization - Evidence of Alcohol/Drug Impairment":  ("Specialty Patient - Spinal Immobilization", "Evidence of Alcohol/Drug Impairment",  "concat_newline"),
    "Spinal Immobilization - Distracting Injury":                   ("Specialty Patient - Spinal Immobilization", "Distracting Injury",                   "concat_newline"),
    "Spinal Immobilization - Neurologic Deficit":                   ("Specialty Patient - Spinal Immobilization", "Neurologic Deficit",                   "concat_newline"),
    "Spinal Immobilization - Spinal Pain/Tenderness":               ("Specialty Patient - Spinal Immobilization", "Spinal Pain/Tenderness",               "concat_newline"),

    # ── Incident Details ─────────────────────────────────────────────────────
    "Incident Details - Location Type":             ("Incident Details", "Location Type"),
    "Incident Details - Location":                  ("Incident Details", "Location"),
    "Incident Details - Address":                   ("Incident Details", "Address"),
    "Incident Details - Address 2":                 ("Incident Details", "Address 2"),
    "Incident Details - Mile Marker":               ("Incident Details", "Mile Marker"),
    "Incident Details - City":                      ("Incident Details", "City"),
    "Incident Details - County":                    ("Incident Details", "County"),
    "Incident Details - State":                     ("Incident Details", "State"),
    "Incident Details - Zip":                       ("Incident Details", "Zip"),
    "Incident Details - Country":                   ("Incident Details", "Country"),
    "Incident Details - Medic Unit":                ("Incident Details", "Medic Unit"),
    "Incident Details - Medic Vehicle":             ("Incident Details", "Medic Vehicle"),
    "Incident Details - Run Type":                  ("Incident Details", "Run Type"),
    "Incident Details - Response Mode":             ("Incident Details", "Response Mode"),
    "Incident Details - Response Mode Descriptors": ("Incident Details", "Response Mode Descriptors"),
    "Incident Details - Shift":                     ("Incident Details", "Shift"),
    "Incident Details - Zone":                      ("Incident Details", "Zone"),
    "Incident Details - Level of Service":          ("Incident Details", "Level of Service"),
    "Incident Details - EMD Complaint":             ("Incident Details", "EMD Complaint"),
    "Incident Details - EMD Card Number":           ("Incident Details", "EMD Card Number"),
    "Incident Details - Dispatch Priority":         ("Incident Details", "Dispatch Priority"),

    # ── Destination Details ───────────────────────────────────────────────────
    "Destination Details - Disposition":                               ("Destination Details", "Disposition"),
    "Destination Details - Unit Disposition":                          ("Destination Details", "Unit Disposition"),
    "Destination Details - Patient Evaluation and/or Care Disposition":("Destination Details", "Patient Evaluation and/or Care Disposition"),
    "Destination Details - Crew Disposition":                          ("Destination Details", "Crew Disposition"),
    "Destination Details - Transport Disposition":                     ("Destination Details", "Transport Disposition"),
    "Destination Details - Transport Mode":                            ("Destination Details", "Transport Mode"),
    "Destination Details - Reason for Refusal or Release":             ("Destination Details", "Reason for Refusal or Release"),
    "Destination Details - Transport Mode Descriptors":                ("Destination Details", "Transport Mode Descriptors"),
    "Destination Details - Transport Due To":                          ("Destination Details", "Transport Due To"),
    "Destination Details - Transported To":                            ("Destination Details", "Transported To"),
    "Destination Details - Requested By":                              ("Destination Details", "Requested By"),
    "Destination Details - Transferred To":                            ("Destination Details", "Transferred To"),
    "Destination Details - Transferred Unit":                          ("Destination Details", "Transferred Unit"),
    "Destination Details - Destination":                               ("Destination Details", "Destination"),
    "Destination Details - Department":                                ("Destination Details", "Department"),
    "Destination Details - Address":                                   ("Destination Details", "Address"),
    "Destination Details - Address 2":                                 ("Destination Details", "Address 2"),
    "Destination Details - City":                                      ("Destination Details", "City"),
    "Destination Details - County":                                    ("Destination Details", "County"),
    "Destination Details - State":                                     ("Destination Details", "State"),
    "Destination Details - Zip":                                       ("Destination Details", "Zip"),
    "Destination Details - Country":                                   ("Destination Details", "Country"),
    "Destination Details - Zone":                                      ("Destination Details", "Zone"),
    "Destination Details - Condition at Destination":                  ("Destination Details", "Condition at Destination"),
    # Word appends "#" to these two; JSON keys do not have it
    "Destination Details - State Wristband #":                         ("Destination Details", "State Wristband"),
    "Destination Details - Destination Record #":                      ("Destination Details", "Destination Record"),
    "Destination Details - Trauma Registry ID":                        ("Destination Details", "Trauma Registry ID"),
    "Destination Details - STEMI Registry ID":                         ("Destination Details", "STEMI Registry ID"),
    "Destination Details - Stroke Registry ID":                        ("Destination Details", "Stroke Registry ID"),

    # ── Incident Times ────────────────────────────────────────────────────────
    "Incident Times - PSAP Call":         ("Incident Times", "PSAP Call"),
    "Incident Times - Dispatch Notified": ("Incident Times", "Dispatch Notified"),
    "Incident Times - Call Received":     ("Incident Times", "Call Received"),
    "Incident Times - Dispatched":        ("Incident Times", "Dispatched"),
    "Incident Times - En Route":          ("Incident Times", "En Route"),
    "Incident Times - Staged":            ("Incident Times", "Staged"),
    "Incident Times - Resp on Scene":     ("Incident Times", "Resp on Scene"),
    "Incident Times - On Scene":          ("Incident Times", "On Scene"),
    "Incident Times - At Patient":        ("Incident Times", "At Patient"),
    "Incident Times - Care Transferred":  ("Incident Times", "Care Transferred"),
    "Incident Times - Depart Scene":      ("Incident Times", "Depart Scene"),
    "Incident Times - At Destination":    ("Incident Times", "At Destination"),
    "Incident Times - Pt. Transferred":   ("Incident Times", "Pt. Transferred"),
    "Incident Times - Call Closed":       ("Incident Times", "Call Closed"),
    "Incident Times - In District":       ("Incident Times", "In District"),
    "Incident Times - At Landing Area":   ("Incident Times", "At Landing Area"),

    # ── Insurance Details (no section prefix in Word) ─────────────────────────
    "Insured's Name":                 ("Insurance Details", "Insured's Name"),
    "Relationship":                   ("Insurance Details", "Relationship"),
    "Insured SSN":                    ("Insurance Details", "Insured SSN"),
    "Insured DOB":                    ("Insurance Details", "Insured DOB"),
    "Address1":                       ("Insurance Details", "Address1"),
    "Address2":                       ("Insurance Details", "Address2"),
    "Address3":                       ("Insurance Details", "Address3"),
    "City":                           ("Insurance Details", "City"),
    "State":                          ("Insurance Details", "State"),
    "Zip":                            ("Insurance Details", "Zip"),
    "Country":                        ("Insurance Details", "Country"),
    "Primary Payer":                  ("Insurance Details", "Primary Payer"),
    "Medicare":                       ("Insurance Details", "Medicare"),
    "Medicaid":                       ("Insurance Details", "Medicaid"),
    "Primary Insurance":              ("Insurance Details", "Primary Insurance"),
    "Primary Insurance Policy #":     ("Insurance Details", "Primary Insurance Policy #"),
    "Primary Insurance Group Name":   ("Insurance Details", "Primary Insurance Group Name"),
    "Primary Insurance Group #":      ("Insurance Details", "Primary Insurance Group #"),
    "Secondary Ins":                  ("Insurance Details", "Secondary Ins"),
    "Secondary Insurance Policy #":   ("Insurance Details", "Secondary Insurance Policy #"),
    "Secondary Insurance Group Name": ("Insurance Details", "Secondary Insurance Group Name"),
    "Secondary Insurance Group #":    ("Insurance Details", "Secondary Insurance Group #"),
    "Dispatch Nature":                ("Insurance Details", "Dispatch Nature"),
    "Response Urgency":               ("Insurance Details", "Response Urgency"),
    "Job Related Injury":             ("Insurance Details", "Job Related Injury"),
    "Employer":                       ("Insurance Details", "Employer"),
    "Contact":                        ("Insurance Details", "Contact"),
    "Phone":                          ("Insurance Details", "Phone"),
    "Mileage to Closest Hospital":    ("Insurance Details", "Mileage to Closest Hospital"),

    # ── Mileage ───────────────────────────────────────────────────────────────
    "Mileage - Scene":        ("Mileage", "Scene"),
    "Mileage - Destination":  ("Mileage", "Destination"),
    "Mileage - Loaded Miles": ("Mileage", "Loaded Miles"),
    "Mileage - Start":        ("Mileage", "Start"),
    "Mileage - End":          ("Mileage", "End"),
    "Mileage - Total Miles":  ("Mileage", "Total Miles"),

    # ── Delays (no section prefix in Word) ────────────────────────────────────
    "Dispatch Delays":    ("Delays", "Dispatch Delays"),
    "Response Delays":    ("Delays", "Response Delays"),
    "Scene Delays":       ("Delays", "Scene Delays"),
    "Transport Delays":   ("Delays", "Transport Delays"),
    "Turn Around Delays": ("Delays", "Turn Around Delays"),

    # ── Additional (no section prefix in Word) ────────────────────────────────
    "Additional Agencies": ("Additional", "Additional Agencies"),

    # ── Consumables (Word indexes each item; data is a list) ──────────────────
    "Consumables 1 Description": ("Consumables", "Description", 0),
    "Consumables 1 Qty":         ("Consumables", "Qty",         0),
    "Consumables 2 Description": ("Consumables", "Description", 1),
    "Consumables 2 Qty":         ("Consumables", "Qty",         1),
    "Consumables 3 Description": ("Consumables", "Description", 2),
    "Consumables 3 Qty":         ("Consumables", "Qty",         2),

    # ── Patient Transport Details ──────────────────────────────────────────────
    "Patient Transport Details - How was Patient Moved To Stretcher":  ("Patient Transport Details", "How was Patient Moved To Stretcher"),
    "Patient Transport Details - How was Patient Moved From Ambulance":("Patient Transport Details", "How was Patient Moved From Ambulance"),
    "Patient Transport Details - Condition of Patient at Destination": ("Patient Transport Details", "Condition of Patient at Destination"),
    "Patient Transport Details - How was Patient Moved To Ambulance":  ("Patient Transport Details", "How was Patient Moved To Ambulance"),
    "Patient Transport Details - Patient Position During Transport":   ("Patient Transport Details", "Patient Position During Transport"),

    # ── Transfer Details ──────────────────────────────────────────────────────
    "Transfer Details - PAN":                          ("Transfer Details", "PAN"),
    "Transfer Details - Prior Authorization Code Payer":("Transfer Details", "Prior Authorization Code Payer"),
    "Transfer Details - PCS":                          ("Transfer Details", "PCS"),
    "Transfer Details - Interfacility Transfer or Medical Transport Reason": ("Transfer Details", "Interfacility Transfer or Medical Transport Reason"),
    "Transfer Details - ABN":                          ("Transfer Details", "ABN"),
    "Transfer Details - CMS Service Level":            ("Transfer Details", "CMS Service Level"),
    # Word has a stray ">" prefix on this key
    "Transfer Details - >ICD-9 Code":                  ("Transfer Details", "ICD-9 Code"),
    "Transfer Details - Transport Assessment":         ("Transfer Details", "Transport Assessment"),
    "Transfer Details - Specialty Care Transport Provider": ("Transfer Details", "Specialty Care Transport Provider"),
    "Transfer Details - Transfer Reason":              ("Transfer Details", "Transfer Reason"),
    "Transfer Details - Justification for Transfer":   ("Transfer Details", "Justification for Transfer"),
    "Transfer Details - Other/Services":               ("Transfer Details", "Other/Services"),
    "Transfer Details - Medical Necessity":            ("Transfer Details", "Medical Necessity"),
    "Transfer Details - Sending Physician":            ("Transfer Details", "Sending Physician"),
    "Transfer Details - Sending Record #":             ("Transfer Details", "Sending Record #"),
    "Transfer Details - Receiving Physician":          ("Transfer Details", "Receiving Physician"),
    "Transfer Details - Condition Code":               ("Transfer Details", "Condition Code"),
    "Transfer Details - Condition Code Modifiers":     ("Transfer Details", "Condition Code Modifiers"),

    # ── Patient Refusal (Word has just the section name; data key is "Signed By") ─
    "Patient Refusal": ("Patient Refusal", "Signed By"),
}

# ── Entry point ────────────────────────────────────────────────────────────────

JSON_DIR    = "./JSON"
OUTPUT_PATH = "./output.xlsx"

if __name__ == "__main__":
    json_files = sorted(glob.glob(os.path.join(JSON_DIR, "*.json")))
    if not json_files:
        print(f"No JSON files found in {JSON_DIR!r}. Run save_JSON_format.py first.")
    else:
        rows = []
        for i, json_path in enumerate(json_files, start=1):
            with open(json_path, encoding="utf-8") as f:
                data = json.load(f)
            rows.append(extract_row(data))
            print(f"  Processed {i}/{len(json_files)}: {os.path.basename(json_path)}")

        export_to_excel(rows, OUTPUT_PATH)
